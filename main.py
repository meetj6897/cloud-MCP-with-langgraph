from fastmcp import FastMCP
# from main import app # importing fastapi app which used incoverting fastmcp code tofastapi
import openpyxl

mcp_demo = FastMCP(name="demo")


@mcp_demo.tool()
def add(a: int, b: int) -> int:
    """Add two numbers."""
    return a + b


@mcp_demo.tool()
def greet(name: str) -> str:
    """Return a greeting message."""
    return f"Hello, {name}, i am demo server!"


@mcp_demo.resource("data://config")
def get_config() -> dict:
    """Return static server config."""
    return {"version": "1.0", "server": "demo"}


@mcp_demo.prompt()
def ask_question(topic: str) -> str:
    """Generate a prompt asking about a topic."""
    return f"Can you explain {topic} in simple terms?"

mcp_exl =FastMCP(name='excel')


@mcp_exl.tool()
def create_excel(filename: str, data: list[list]) -> str:
    """Create an Excel file with given data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(filename)
    return f"Created {filename}"

@mcp_exl.tool()
def update_cell(filename: str, row: int, col: int, value: str) -> str:
    """Update a specific cell in an Excel file."""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    ws.cell(row=row, column=col, value=value)
    wb.save(filename)
    return f"Updated cell ({row},{col}) to {value}"

@mcp_exl.tool()
def read_excel(filename: str) -> list:
    """Read all data from an Excel file."""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]



mcp=FastMCP(name="main")
# mcp=FastMCP.from_fastapi(app=app,name='give name') # to convert he fastapi app to mcp server
mcp.mount(mcp_demo, 'demo')
mcp.mount(mcp_exl, 'excel')

if __name__ == "__main__":
    mcp.run(transport="http",host="0.0.0.0",port=8000) # to run on mcp cloud in local it default stdio method
#now to deploy the server we do fastmcp cloud make account in it ans push this code into github